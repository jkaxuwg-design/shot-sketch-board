from __future__ import annotations

import argparse
import base64
import csv
import html
import re
import zipfile
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def safe_name(value: object, max_len: int = 34) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"[\\/:*?\"<>|\r\n\t]+", "-", text).strip()
    text = re.sub(r"\s+", "", text)
    return text[:max_len] or "unnamed-shot"


def image_extension(data: bytes) -> str:
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if data.startswith(b"\xff\xd8"):
        return ".jpg"
    if data.startswith(b"GIF87a") or data.startswith(b"GIF89a"):
        return ".gif"
    return ".bin"


def image_mime(data: bytes) -> str:
    ext = image_extension(data)
    if ext == ".jpg":
        return "image/jpeg"
    if ext == ".gif":
        return "image/gif"
    return "image/png"


def data_uri(data: bytes) -> str:
    return f"data:{image_mime(data)};base64,{base64.b64encode(data).decode('ascii')}"


def find_header_row(ws) -> int:
    for row in range(1, min(ws.max_row, 40) + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        if "镜头/段落/时长" in values and any("草图" in str(v) for v in values if v is not None):
            return row
    raise RuntimeError("Cannot find storyboard header row with 镜头/段落/时长 and 草图 columns.")


def find_column(ws, header_row: int, header_text: str) -> int | None:
    for col in range(1, ws.max_column + 1):
        if ws.cell(header_row, col).value == header_text:
            return col
    return None


def find_sketch_image_column(ws, header_row: int) -> int:
    preferred = {"草图/画面", "草图/画面备注"}
    for col in range(1, ws.max_column + 1):
        if ws.cell(header_row, col).value in preferred:
            return col
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(header_row, col).value or "")
        if "草图" in header and header not in {"草图文件", "飞书复制说明"}:
            return col
    raise RuntimeError("Cannot find sketch image column.")


def copy_header_style(src_cell, dst_cell) -> None:
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)
    dst_cell.font = Font(bold=True, color="000000")
    dst_cell.fill = PatternFill("solid", fgColor="EAF2C2")
    dst_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    dst_cell.border = Border(
        left=Side(style="thin", color="333333"),
        right=Side(style="thin", color="333333"),
        top=Side(style="thin", color="333333"),
        bottom=Side(style="thin", color="333333"),
    )


def cell_text(value: object) -> str:
    if value is None:
        return ""
    text = re.sub(r"\r\n?", "\n", str(value))
    return html.escape(text).replace("\n", "<br>")


def html_css() -> str:
    return """
body{font-family:Arial,"Microsoft YaHei",sans-serif;margin:24px;color:#111827;background:#fff}
.note{margin:0 0 16px;color:#4b5563;font-size:13px}
table{border-collapse:collapse;table-layout:fixed;width:max-content;max-width:none}
td,th{border:1px solid #222;padding:8px;vertical-align:middle;font-size:12px;line-height:1.45;white-space:normal;word-break:break-word;background:#fff}
th{background:#eef7c8;text-align:center;font-weight:700}
.meta-label{background:#dbe6fb;font-weight:700;text-align:center}
.sketch-cell{background:#fff}
.sketch-img{display:block;width:560px;max-width:560px;height:auto;margin:0 auto}
.sketch-caption{font-size:11px;color:#64748b;margin-bottom:6px}
.narrow{width:90px}.medium{width:180px}.wide{width:320px}.script{width:420px}.prompt{width:360px}.sketch{width:600px}
"""


def col_class(col: int) -> str:
    if col in (1, 2, 5, 6):
        return "narrow"
    if col in (7, 8, 14):
        return "medium"
    if col == 4:
        return "script"
    if col in (9, 10, 11, 12, 15):
        return "prompt"
    if col == 13:
        return "sketch"
    return "wide"


def wrap_html(title: str, table_rows: str) -> str:
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>{html.escape(title)}</title>
<style>{html_css()}</style>
</head>
<body>
<p class="note">飞书粘贴专用版：草图是表格单元格里的内联图片，不是 Excel 浮动图片对象。若粘贴平台拦截 data:image 图片，请使用图包或飞书 API 上传图片。</p>
<table>
{table_rows}
</table>
</body>
</html>
"""


def sketch_label(row: int, shot: object) -> str:
    return html.escape(f"ROW {row} / {'' if shot is None else shot}")


def collect_images(src: Path, sheet_name: str | None, image_dir: Path) -> tuple[str, dict[int, list[dict[str, str]]]]:
    wb = load_workbook(src)
    ws = wb[sheet_name] if sheet_name else wb.active
    actual_sheet_name = ws.title
    image_dir.mkdir(parents=True, exist_ok=True)
    for old in image_dir.glob("*"):
        if old.is_file():
            old.unlink()

    images_by_row: dict[int, list[dict[str, str]]] = {}
    for idx, img in enumerate(ws._images, start=1):
        row = img.anchor._from.row + 1
        data = img._data()
        ext = image_extension(data)
        shot = ws.cell(row, 1).value
        filename = f"SHOT_{idx:03d}_ROW{row:03d}_{safe_name(shot)}{ext}"
        path = image_dir / filename
        path.write_bytes(data)

        description = ws.cell(row, 3).value or ""
        line = ws.cell(row, 4).value or ""
        meaning_source = str(description or line).replace("\n", " ").strip()
        if len(meaning_source) > 110:
            meaning_source = meaning_source[:110] + "..."
        meaning = f"对应草图：{filename}。复制到飞书后如图片丢失，从图包插入此文件；画面含义：{meaning_source}"

        images_by_row.setdefault(row, []).append(
            {
                "filename": filename,
                "path": str(path),
                "data_uri": data_uri(data),
                "meaning": meaning,
                "shot": str(shot or ""),
                "line": str(line or "").replace("\n", " "),
            }
        )

    wb.close()
    return actual_sheet_name, images_by_row


def write_feishu_workbook(
    src: Path,
    out_xlsx: Path,
    sheet_name: str,
    image_dir_name: str,
    images_by_row: dict[int, list[dict[str, str]]],
) -> None:
    wb = load_workbook(src)
    ws = wb[sheet_name]
    header_row = find_header_row(ws)

    file_col = find_column(ws, header_row, "草图文件") or ws.max_column + 1
    note_col = find_column(ws, header_row, "飞书复制说明") or max(ws.max_column + 1, file_col + 1)
    source_header = ws.cell(header_row, max(1, file_col - 1))

    ws.cell(header_row, file_col).value = "草图文件"
    ws.cell(header_row, note_col).value = "飞书复制说明"
    copy_header_style(source_header, ws.cell(header_row, file_col))
    copy_header_style(source_header, ws.cell(header_row, note_col))
    ws.column_dimensions[ws.cell(header_row, file_col).column_letter].width = 38
    ws.column_dimensions[ws.cell(header_row, note_col).column_letter].width = 54

    body_alignment = Alignment(vertical="center", wrap_text=True)
    for row in range(header_row + 1, ws.max_row + 1):
        entries = images_by_row.get(row)
        if entries:
            ws.cell(row, file_col).value = "\n".join(entry["filename"] for entry in entries)
            ws.cell(row, note_col).value = "\n".join(entry["meaning"] for entry in entries)
            if len(entries) == 1:
                ws.cell(row, file_col).hyperlink = entries[0]["path"]
                ws.cell(row, file_col).style = "Hyperlink"
        else:
            ws.cell(row, file_col).value = None
            ws.cell(row, note_col).value = None
        ws.cell(row, file_col).alignment = body_alignment
        ws.cell(row, note_col).alignment = body_alignment

    ws.cell(5, file_col).value = "飞书提示：Excel 图片通常是浮动对象，复制到飞书可能丢失；请保留本列文件名并从图包插入。"
    ws.cell(5, note_col).value = f"草图图包：{image_dir_name}"
    ws.cell(5, file_col).alignment = body_alignment
    ws.cell(5, note_col).alignment = body_alignment
    wb.save(out_xlsx)
    wb.close()


def write_index(index_tsv: Path, images_by_row: dict[int, list[dict[str, str]]]) -> None:
    with index_tsv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerow(["Excel行号", "镜头/段落/时长", "台词/旁白", "草图文件", "飞书复制说明"])
        for row in sorted(images_by_row):
            for entry in images_by_row[row]:
                writer.writerow([row, entry["shot"], entry["line"], entry["filename"], entry["meaning"]])


def render_full_html(src: Path, sheet_name: str, images_by_row: dict[int, list[dict[str, str]]]) -> str:
    wb = load_workbook(src)
    ws = wb[sheet_name]
    header_row = find_header_row(ws)
    sketch_col = find_sketch_image_column(ws, header_row)
    rows: list[str] = []
    for row in range(1, ws.max_row + 1):
        tag = "th" if row == header_row else "td"
        cells: list[str] = []
        for col in range(1, ws.max_column + 1):
            classes = [col_class(col)]
            if row <= 5 and col == 1:
                classes.append("meta-label")
            content = cell_text(ws.cell(row, col).value)
            if col == sketch_col and row in images_by_row:
                classes.append("sketch-cell")
                parts = [f'<div class="sketch-caption">{sketch_label(row, ws.cell(row, 1).value)}</div>']
                parts.extend(f'<img class="sketch-img" src="{entry["data_uri"]}" alt="shot sketch row {row}">' for entry in images_by_row[row])
                content = "".join(parts)
            cells.append(f'<{tag} class="{" ".join(classes)}">{content}</{tag}>')
        rows.append("<tr>" + "".join(cells) + "</tr>")
    wb.close()
    return "\n".join(rows)


def render_sketch_only_html(src: Path, sheet_name: str, images_by_row: dict[int, list[dict[str, str]]]) -> str:
    wb = load_workbook(src)
    ws = wb[sheet_name]
    rows = [
        '<tr><th class="narrow">Excel行号</th><th class="narrow">镜头/段落/时长</th><th class="script">台词/旁白</th><th class="sketch">草图/画面</th></tr>'
    ]
    for row in sorted(images_by_row):
        image_parts = [f'<div class="sketch-caption">{sketch_label(row, ws.cell(row, 1).value)}</div>']
        image_parts.extend(f'<img class="sketch-img" src="{entry["data_uri"]}" alt="shot sketch row {row}">' for entry in images_by_row[row])
        rows.append(
            "<tr>"
            f'<td class="narrow">{row}</td>'
            f'<td class="narrow">{cell_text(ws.cell(row, 1).value)}</td>'
            f'<td class="script">{cell_text(ws.cell(row, 4).value)}</td>'
            f'<td class="sketch sketch-cell">{"".join(image_parts)}</td>'
            "</tr>"
        )
    wb.close()
    return "\n".join(rows)


def zip_images(image_dir: Path, zip_path: Path) -> None:
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for path in sorted(image_dir.glob("*")):
            if path.is_file():
                z.write(path, path.name)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create Feishu-safe handoff files for storyboard workbooks with embedded sketch images.")
    parser.add_argument("xlsx", type=Path, help="Input storyboard .xlsx file")
    parser.add_argument("--output-dir", type=Path, default=None, help="Output directory. Default: beside input workbook")
    parser.add_argument("--sheet-name", default=None, help="Worksheet name. Default: active sheet")
    parser.add_argument("--prefix", default=None, help="Output file prefix. Default: input file stem")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    src = args.xlsx.resolve()
    out_dir = (args.output_dir or src.parent).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    prefix = args.prefix or src.stem

    image_dir = out_dir / f"{prefix}_草图图包_飞书迁移版"
    sheet_name, images_by_row = collect_images(src, args.sheet_name, image_dir)

    out_xlsx = out_dir / f"{prefix}_飞书迁移版.xlsx"
    index_tsv = out_dir / f"{prefix}_草图索引_飞书复制.tsv"
    full_html = out_dir / f"{prefix}_飞书粘贴版_图片在单元格.html"
    sketch_only_html = out_dir / f"{prefix}_飞书草图列粘贴版.html"
    zip_path = out_dir / f"{prefix}_草图图包_飞书迁移版.zip"

    write_feishu_workbook(src, out_xlsx, sheet_name, image_dir.name, images_by_row)
    write_index(index_tsv, images_by_row)
    zip_images(image_dir, zip_path)
    full_html.write_text(wrap_html(f"{prefix} 飞书粘贴版", render_full_html(out_xlsx, sheet_name, images_by_row)), encoding="utf-8")
    sketch_only_html.write_text(wrap_html(f"{prefix} 飞书草图列粘贴版", render_sketch_only_html(out_xlsx, sheet_name, images_by_row)), encoding="utf-8")

    print(f"input={src}")
    print(f"sheet={sheet_name}")
    print(f"workbook={out_xlsx}")
    print(f"image_dir={image_dir}")
    print(f"image_zip={zip_path}")
    print(f"index_tsv={index_tsv}")
    print(f"full_html={full_html}")
    print(f"sketch_only_html={sketch_only_html}")
    print(f"rows_with_images={len(images_by_row)}")
    print(f"images={sum(len(v) for v in images_by_row.values())}")


if __name__ == "__main__":
    main()
